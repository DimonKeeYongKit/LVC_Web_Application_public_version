from openpyxl import load_workbook, Workbook

filePath = ""

def create_crawler_workbook(book_name, book_type, platform):
    book_name = book_name.lower()
    if book_type != "Comment" and book_type != "Post":
        print("Error workbook type!")
        return False
    wb = Workbook()
    wbs = wb.active
    if platform == 'Twitter':
        wbs.append(['post_id', 'post_content', 'post_date', 'post_time', 'post_url', 'author_id', 'author_name',
                    'crawl_date', 'crawl_time'])
    elif platform == 'Facebook':
        wbs.append(['post_id', 'post_content', 'post_date', 'post_time', 'post_url', 'author_id', 'author_name',
                    'crawl_date', 'crawl_time'])
    elif platform == 'Reddit':
        wbs.append(['post_id', 'post_title', 'post_content', 'post_date', 'post_time', 'post_url', 'author_id',
                    'author_name', 'crawl_date', 'crawl_time'])
    else:
        wb.close()
        print('Error platform selection!')
        return False

    # Save the data
    wb.save(filePath + platform + "/" + book_type + "/" + book_name + '.xlsx')
    wb.close()
    return True


def load_crawler_workbook(book_name, book_type, platform):
    book_name = book_name.lower()
    content_list = []

    if book_type != "Comment" and book_type != "Post":
        print("Error workbook type!")
        return False

    if platform != "Facebook" and platform != "Reddit" and platform != "Twitter":
        print("Error platform type!")
        return False

    try:
        wb = load_workbook(filePath + platform + "/" + book_type + "/" + book_name + ".xlsx")
        wbs = wb.active
        max_row = wbs.max_row
    except FileNotFoundError:
        return False

    for i in range(2, max_row + 1):
        cell_obj = wbs.cell(row=i, column=2)
        content_list.append(cell_obj.value)

    wb.close()
    return content_list


def load_full_workbook(book_name, book_type, platform):
    book_name = book_name.lower()
    content_list = []

    if book_type != "Comment" and book_type != "Post":
        print("Error workbook type!")
        return False

    if platform != "Facebook" and platform != "Reddit" and platform != "Twitter":
        print("Error platform type!")
        return False

    try:
        wb = load_workbook(filePath + platform + "/" + book_type + "/" + book_name + ".xlsx")
        wbs = wb.active
    except FileNotFoundError:
        return False

    for row in wbs.iter_rows(max_row=wbs.max_row, max_col=wbs.max_column, min_row=2):
        current_row = []
        for cell in row:
            current_row.append(cell.value)
        content_list.append(current_row)

    wb.close()
    return content_list


def load_lite_workbook(book_name, book_type, platform):
    book_name = book_name.lower()
    content_list = []

    if book_type != "Comment" and book_type != "Post":
        print("Error workbook type!")
        return False

    if platform != "Facebook" and platform != "Reddit" and platform != "Twitter":
        print("Error platform type!")
        return False

    try:
        wb = load_workbook(filePath + platform + "/" + book_type + "/" + book_name + ".xlsx")
        wbs = wb.active
        max_row = wbs.max_row
    except FileNotFoundError:
        return False

    for i in range(2, max_row + 1):
        id = wbs.cell(row=i, column=1)
        content = wbs.cell(row=i, column=2)
        if platform == 'Facebook' or platform == 'Twitter':
            post_date = wbs.cell(row=i, column=3)
            post_time = wbs.cell(row=i, column=4)
        elif platform == 'Reddit':
            post_date = wbs.cell(row=i, column=4)
            post_time = wbs.cell(row=i, column=5)
        content_list.append([id.value, content.value, post_date.value, post_time.value])

    wb.close()
    return content_list


def check_crawler_workbook(book_name, book_type, platform):
    book_name = book_name.lower()
    if book_type != "Comment" and book_type != "Post":
        print("Error workbook type!")
        return False

    if platform != "Facebook" and platform != "Reddit" and platform != "Twitter":
        print("Error platform type!")
        return False

    try:
        f = open(filePath + platform + "/" + book_type + "/" + book_name + ".xlsx")
        f.close()
        return True
    except IOError:
        return False


def add_row_workbook(book_name, book_type, platform, dataset):
    book_name = book_name.lower()
    if book_type != "Comment" and book_type != "Post":
        print("Error workbook type!")
        return False

    if platform != "Facebook" and platform != "Reddit" and platform != "Twitter":
        print("Error platform type!")
        return False

    try:
        book = filePath + platform + "/" + book_type + "/" + book_name + ".xlsx"
        wb = load_workbook(book)
        wbs = wb.active
    except FileNotFoundError:
        return False

    for data in dataset:
        wbs.append(data)

    wb.save(book)
    wb.close()
    return True


def delete_row_workbook(book_name, book_type, platform, row):
    book_name = book_name.lower()
    if book_type != "Comment" and book_type != "Post":
        print("Error workbook type!")
        return False

    if platform != "Facebook" and platform != "Reddit" and platform != "Twitter":
        print("Error platform type!")
        return False
    try:
        book = filePath + platform + "/" + book_type + "/" + book_name + ".xlsx"
        wb = load_workbook(book)
        wbs = wb.active
    except FileNotFoundError:
        return False

    # check valid input
    if wbs.max_row < row:
        wb.close()
        return False
    else:
        wbs.delete_rows(row, 1)  # first para(row) is the start row,second(1) para is how much row
        wb.save(book)
        wb.close()
        return True


def load_exist_postID(book_name, book_type, platform):
    book_name = book_name.lower()
    exist_post_id = []
    if book_type != "Comment" and book_type != "Post":
        print("Error workbook type!")
        return False

    if platform != "Facebook" and platform != "Reddit" and platform != "Twitter":
        print("Error platform type!")
        return False
    try:
        book = filePath + platform + "/" + book_type + "/" + book_name + ".xlsx"
        wb = load_workbook(book)
        wbs = wb.active
        max_row = wbs.max_row
    except FileNotFoundError:
        return False

    # Loop will load the id
    for i in range(2, max_row + 1):
        cell_obj = wbs.cell(row=i, column=1)
        exist_post_id.append(cell_obj.value)

    wb.close()
    return exist_post_id


def save_crawler_workbook(book_name, book_type, platform, dataset):
    book_name = book_name.lower()
    exist = check_crawler_workbook(book_name, book_type, platform)
    if not exist:
        create_crawler_workbook(book_name, book_type, platform)
    add_row_workbook(book_name, book_type, platform, dataset)
